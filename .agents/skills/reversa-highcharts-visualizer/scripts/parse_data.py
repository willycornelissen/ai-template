#!/usr/bin/env python3
"""
Parseia dados de CSV, JSON ou Excel e formata para uso em Highcharts.

Detecta automaticamente o formato, encoding, e estrutura dos dados.
Saída: JSON pronto para ser embutido nas opções do Highcharts.

Uso:
    python parse_data.py <arquivo> [--format categories|timeseries|xy|pie]
    python parse_data.py dados.csv --sheet "Plan1" --encoding utf-8
    python parse_data.py dados.json --output formatted.json

Saída:
    JSON com: { categories, series, metadata }
"""

import sys
import json
import argparse
from pathlib import Path


def detect_encoding(filepath: str) -> str:
    """Tenta detectar encoding do arquivo."""
    encodings = ['utf-8', 'utf-8-sig', 'latin1', 'iso-8859-1', 'cp1252']
    for enc in encodings:
        try:
            with open(filepath, 'r', encoding=enc) as f:
                f.read(1000)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return 'utf-8'


def parse_number(value: str) -> float | None:
    """Converte string para número, tratando formatos BR e US."""
    if not value or not isinstance(value, str):
        return value if isinstance(value, (int, float)) else None
    value = value.strip().replace(' ', '')
    # Formato BR: 1.234,56
    if ',' in value and '.' in value and value.rindex(',') > value.rindex('.'):
        value = value.replace('.', '').replace(',', '.')
    # Formato BR sem milhar: 123,45
    elif ',' in value and '.' not in value:
        value = value.replace(',', '.')
    # Remover símbolos de moeda
    for symbol in ['R$', '$', '€', '£', '%']:
        value = value.replace(symbol, '')
    try:
        return float(value)
    except ValueError:
        return None


def parse_csv(filepath: str, encoding: str = 'utf-8', delimiter: str = None) -> dict:
    """Parseia CSV para formato Highcharts."""
    import csv

    with open(filepath, 'r', encoding=encoding) as f:
        content = f.read()

    # Detectar delimitador
    if delimiter is None:
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(content[:2000])
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ',' if ',' in content else ';' if ';' in content else '\t'

    lines = content.strip().split('\n')
    reader = csv.reader(lines, delimiter=delimiter)
    rows = list(reader)

    if len(rows) < 2:
        return {"error": "Arquivo com menos de 2 linhas"}

    headers = [h.strip() for h in rows[0]]
    data_rows = rows[1:]

    # Primeira coluna = categorias, demais = séries
    categories = [row[0].strip() for row in data_rows if row]
    series = []
    for col_idx in range(1, len(headers)):
        values = []
        for row in data_rows:
            if col_idx < len(row):
                val = parse_number(row[col_idx])
                values.append(val if val is not None else 0)
            else:
                values.append(0)
        series.append({
            "name": headers[col_idx],
            "data": values
        })

    return {
        "categories": categories,
        "series": series,
        "metadata": {
            "rows": len(data_rows),
            "columns": len(headers),
            "headers": headers,
            "delimiter": delimiter,
            "encoding": encoding
        }
    }


def parse_json_data(filepath: str) -> dict:
    """Parseia JSON para formato Highcharts."""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Se já é formato Highcharts, retornar direto
    if isinstance(data, dict) and 'series' in data:
        return data

    # Se é array de objetos [{x: ..., y: ...}]
    if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
        keys = list(data[0].keys())
        category_key = keys[0]
        categories = [str(item.get(category_key, '')) for item in data]
        series = []
        for key in keys[1:]:
            values = [parse_number(str(item.get(key, 0))) or 0 for item in data]
            series.append({"name": key, "data": values})
        return {
            "categories": categories,
            "series": series,
            "metadata": {"rows": len(data), "keys": keys, "format": "array_of_objects"}
        }

    # Se é array de arrays [[cat, v1, v2], ...]
    if isinstance(data, list) and len(data) > 0 and isinstance(data[0], list):
        categories = [str(row[0]) for row in data[1:]]
        headers = data[0]
        series = []
        for col_idx in range(1, len(headers)):
            values = [parse_number(str(row[col_idx])) or 0
                      for row in data[1:] if col_idx < len(row)]
            series.append({"name": str(headers[col_idx]), "data": values})
        return {
            "categories": categories,
            "series": series,
            "metadata": {"rows": len(data) - 1, "format": "array_of_arrays"}
        }

    return {"error": "Formato JSON não reconhecido", "raw": data}


def parse_excel(filepath: str, sheet: str = None) -> dict:
    """Parseia Excel para formato Highcharts."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)

    if sheet:
        ws = wb[sheet]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"error": "Planilha com menos de 2 linhas"}

    headers = [str(h).strip() if h else f"Col_{i}" for i, h in enumerate(rows[0])]
    data_rows = rows[1:]

    categories = [str(row[0]).strip() if row[0] else '' for row in data_rows]
    series = []
    for col_idx in range(1, len(headers)):
        values = []
        for row in data_rows:
            val = row[col_idx] if col_idx < len(row) else 0
            if isinstance(val, (int, float)):
                values.append(val)
            else:
                parsed = parse_number(str(val)) if val else 0
                values.append(parsed or 0)
        series.append({"name": headers[col_idx], "data": values})

    return {
        "categories": categories,
        "series": series,
        "metadata": {
            "rows": len(data_rows),
            "columns": len(headers),
            "headers": headers,
            "sheet": ws.title,
            "sheets_available": wb.sheetnames
        }
    }


def main():
    parser = argparse.ArgumentParser(description="Parseia dados para formato Highcharts")
    parser.add_argument("filepath", help="Caminho do arquivo de dados")
    parser.add_argument("--encoding", default=None, help="Encoding do arquivo CSV")
    parser.add_argument("--delimiter", default=None, help="Delimitador CSV")
    parser.add_argument("--sheet", default=None, help="Nome da aba Excel")
    parser.add_argument("--output", "-o", help="Salvar resultado em arquivo")
    args = parser.parse_args()

    path = Path(args.filepath)
    if not path.exists():
        print(f"[ERRO] Arquivo não encontrado: {path}", file=sys.stderr)
        sys.exit(1)

    ext = path.suffix.lower()

    if ext in ('.csv', '.tsv', '.txt'):
        encoding = args.encoding or detect_encoding(str(path))
        result = parse_csv(str(path), encoding=encoding, delimiter=args.delimiter)
    elif ext == '.json':
        result = parse_json_data(str(path))
    elif ext in ('.xlsx', '.xls'):
        result = parse_excel(str(path), sheet=args.sheet)
    else:
        print(f"[ERRO] Formato não suportado: {ext}", file=sys.stderr)
        sys.exit(1)

    if "error" in result:
        print(f"[ERRO] {result['error']}", file=sys.stderr)
        sys.exit(1)

    meta = result.get("metadata", {})
    print(f"[INFO] Linhas: {meta.get('rows', '?')}, "
          f"Colunas: {meta.get('columns', len(result.get('series', [])) + 1)}", file=sys.stderr)
    if 'series' in result:
        for s in result['series']:
            print(f"[INFO]   Série '{s['name']}': {len(s['data'])} pontos", file=sys.stderr)

    output = json.dumps(result, ensure_ascii=False, indent=2)

    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(output)
        print(f"[INFO] Salvo em: {args.output}", file=sys.stderr)
    else:
        print(output)


if __name__ == "__main__":
    main()
